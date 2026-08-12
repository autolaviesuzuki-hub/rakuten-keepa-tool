import json
import openpyxl

# ===============================
# results.json 読み込み
# ===============================
with open("../output/results.json", "r", encoding="utf-8") as f:
    data = json.load(f)

# ===============================
# Excelテンプレート読み込み
# ===============================
wb = openpyxl.load_workbook("../RakutenPointTemplate.xlsx")
ws = wb.active

start_row = 12  # あなたの既存ツールと同じ
row = start_row

# ===============================
# results.json → Excel 書き込み
# ===============================
for item in data:
    ws[f"A{row}"] = item.get("asin")
    ws[f"B{row}"] = item.get("model")
    ws[f"C{row}"] = item.get("shop")
    ws[f"D{row}"] = item.get("price")
    ws[f"E{row}"] = item.get("url")
    ws[f"F{row}"] = item.get("matchedModel")
    row += 1

# ===============================
# Excel保存
# ===============================
wb.save("../RakutenPointTemplate_output.xlsx")
print("Excel 出力完了")
